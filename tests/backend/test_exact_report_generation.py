import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2] / "backend"
sys.path.insert(0, str(ROOT))

import openpyxl
import pandas as pd
import pytest

from app.reconciliation.generic import run_generic_reconciliation
from app.reconciliation.gst import run_gst_reconciliation
from app.reconciliation_engine.universal_mapper import build_universal_data_model
from app.reconciliation_engine.universal_reporter import generate_enterprise_report


def test_generated_excel_structure_matches_sample(tmp_path: Path):
    output_path = tmp_path / "generated_report.xlsx"

    # Sample reconciliation data model
    data = {
        "metadata": {
            "reconciliation_name": "Generic Reconciliation",
            "file_1_name": "Books.xlsx",
            "file_2_name": "Portal.xlsx",
            "matching_keys": ["EMP#"],
        },
        "statistics": {
            "total_file_1": 551,
            "total_file_2": 551,
            "matched": 542,
            "mismatched": 9,
            "missing_in_file_1": 10,
            "missing_in_file_2": 10,
        },
        "overall_status": "EXCEPTIONS FOUND",
        "exceptions": [
            {
                "Exception ID": "EX-000001",
                "Match Key": "IGS0523",
                "Field": "TOTAL ST",
                "File 1 Value": 33,
                "File 2 Value": 32,
                "Difference": 1.0,
                "Difference %": 0.03125,
                "Exception Type": "Value Difference",
                "Severity": "Medium",
                "Status": "Open",
                "Action Notes": "",
            },
            {
                "Exception ID": "EX-000002",
                "Match Key": "IGS0259",
                "Field": "TOTAL OT",
                "File 1 Value": 3,
                "File 2 Value": 2.2,
                "Difference": 0.8,
                "Difference %": 0.3636,
                "Exception Type": "Value Difference",
                "Severity": "Medium",
                "Status": "Open",
                "Action Notes": "",
            },
            {
                "Exception ID": "EX-000003",
                "Match Key": "IGS1324",
                "Field": "EMPLOYEE - FIRST",
                "File 1 Value": "Deborah",
                "File 2 Value": "geller",
                "Difference": None,
                "Difference %": None,
                "Exception Type": "Text Difference",
                "Severity": "Medium",
                "Status": "Open",
                "Action Notes": "",
            },
        ],
        "matched_records": [
            {
                "ROW (FILE 1)": 11,
                "ROW (FILE 2)": 11,
                "EMP#": "IGS1344",
                "MATCHED EMP#": "IGS1344",
                "MATCH TYPE": "numeric",
                "MATCH CONFIDENCE": "100%",
                "MATCH STATUS": "Exact numeric match",
            }
        ],
        "missing_in_file_1": [
            {
                "EMP#": "IGS09091",
                "Client": "Facebook",
                "Employee - Last": "Saransh",
                "Employee - First": "Khajuria",
                "Total ST": 72,
                "portal.xlsx Row": 586,
            }
        ],
        "missing_in_file_2": [
            {
                "EMP#": "IGS090945",
                "Client": "Google",
                "Employee - Last": "Jagdev",
                "Employee - First": "Singh",
                "Total ST": 40,
                "Books.xlsx Row": 586,
            }
        ],
        "field_differences": [
            {
                "Match Key": "IGS0523",
                "Field": "TOTAL ST",
                "File 1 Value": 33,
                "File 2 Value": 32,
                "Difference": 1.0,
                "Difference %": 0.03125,
                "Result": "Mismatch",
            }
        ],
        "field_exception_summary": [
            {"Field": "TOTAL ST", "Matched": 542, "Mismatch": 6, "Match %": 0.989},
            {"Field": "EMPLOYEE - FIRST", "Matched": 542, "Mismatch": 4, "Match %": 0.993},
            {"Field": "TOTAL OT", "Matched": 542, "Mismatch": 1, "Match %": 0.998},
        ],
        "control_checks": [],
    }

    generate_enterprise_report(data, {}, output_path)

    assert output_path.exists()
    wb = openpyxl.load_workbook(output_path, data_only=False)

    # 1. Verify sheet names and count
    expected_sheets = [
        "01 Executive Summary",
        "02 Exceptions",
        "03 Matched Records",
        "04 Missing - Books",
        "05 Missing - Portal",
        "06 Field Differences",
        "07 Control Checks",
    ]
    assert wb.sheetnames == expected_sheets

    # 2. Verify Tab Colors
    assert wb["01 Executive Summary"].sheet_properties.tabColor.rgb in ["FF1F3864", "001F3864", "1F3864"]
    assert wb["02 Exceptions"].sheet_properties.tabColor.rgb in ["FFC0392B", "00C0392B", "C0392B"]
    assert wb["03 Matched Records"].sheet_properties.tabColor.rgb in ["FF1E7B34", "001E7B34", "1E7B34"]
    assert wb["04 Missing - Books"].sheet_properties.tabColor.rgb in ["FFA6192E", "00A6192E", "A6192E"]
    assert wb["05 Missing - Portal"].sheet_properties.tabColor.rgb in ["FFA6192E", "00A6192E", "A6192E"]
    assert wb["06 Field Differences"].sheet_properties.tabColor.rgb in ["FFB36A00", "00B36A00", "B36A00"]
    assert wb["07 Control Checks"].sheet_properties.tabColor.rgb in ["FF1F3864", "001F3864", "1F3864"]

    # 3. Verify showGridLines = False on all sheets
    for name in expected_sheets:
        ws = wb[name]
        assert ws.views.sheetView[0].showGridLines is False

    # 4. Verify 01 Executive Summary Content & Formulas
    ws1 = wb["01 Executive Summary"]
    assert ws1["B1"].value == "RECLIQ  |  RECONCILIATION EXECUTIVE SUMMARY"
    assert "Books.xlsx" in ws1["B2"].value
    assert "Portal.xlsx" in ws1["B2"].value
    assert "EMP#" in ws1["B2"].value

    # KPI formula checks
    assert "03 Matched Records" in ws1["B6"].value
    assert "02 Exceptions" in ws1["B6"].value
    assert ws1["J6"].value == "=IF($B$6>0,$D$6/$B$6,0)"
    assert ws1["L6"].value == "=IF($B$6>0,$F$6/$B$6,0)"

    # Charts on 01 Executive Summary
    assert len(ws1._charts) == 2
    chart_titles = []
    for c in ws1._charts:
        if isinstance(c.title, str):
            chart_titles.append(c.title)
        elif hasattr(c.title, "tx") and c.title.tx and hasattr(c.title.tx, "rich") and c.title.tx.rich:
            paragraphs = c.title.tx.rich.p
            text = "".join(r.t for p in paragraphs for r in p.r if hasattr(r, "t") and r.t)
            chart_titles.append(text)
        else:
            chart_titles.append(str(c.title))
    assert "Reconciliation Outcome" in chart_titles
    assert "Field-Level Mismatches" in chart_titles

    # 5. Verify 02 Exceptions Distinct Record formula
    ws2 = wb["02 Exceptions"]
    assert ws2["L4"].value == "Distinct Record"
    assert ws2["L5"].value == "=IF(COUNTIF($B$5:B5,B5)=1,1,0)"
    assert ws2.freeze_panes == "A5"

    # 6. Verify 03 Matched Records Table
    ws3 = wb["03 Matched Records"]
    assert "MatchedRecordsTbl" in [t.displayName for t in ws3.tables.values()]

    # 7. Verify 07 Control Checks formulas and conditional formatting
    ws7 = wb["07 Control Checks"]
    assert ws7["D5"].value == '=IF(B5=C5,"Pass","Review")'
    assert len(ws7.conditional_formatting) > 0


def test_generic_reconciliation_end_to_end(tmp_path: Path):
    file1 = tmp_path / "ledger.xlsx"
    file2 = tmp_path / "bank.xlsx"
    output = tmp_path / "Reconciliation_Report.xlsx"

    pd.DataFrame([
        {"EmpId": "E101", "Name": "Alice Smith", "Hours": 40},
        {"EmpId": "E102", "Name": "Bob Jones", "Hours": 35},
        {"EmpId": "E103", "Name": "Charlie Brown", "Hours": 20},
    ]).to_excel(file1, index=False)

    pd.DataFrame([
        {"EmpId": "E101", "Name": "Alice Smith", "Hours": 40},
        {"EmpId": "E102", "Name": "Bob Jones", "Hours": 38}, # Mismatched hours
        {"EmpId": "E104", "Name": "David Clark", "Hours": 25}, # Missing in file 1
    ]).to_excel(file2, index=False)

    summary = run_generic_reconciliation(
        file1,
        file2,
        output,
        key_file_1="EmpId",
        key_file_2="EmpId",
        rules=[
            {"file_1_fields": ["Hours"], "file_2_fields": ["Hours"]},
            {"file_1_fields": ["Name"], "file_2_fields": ["Name"]},
        ],
        file_1_name="ledger.xlsx",
        file_2_name="bank.xlsx",
    )

    assert output.exists()
    wb = openpyxl.load_workbook(output)
    assert "01 Executive Summary" in wb.sheetnames
    assert "02 Exceptions" in wb.sheetnames
    assert "03 Matched Records" in wb.sheetnames
    assert "04 Missing - ledger" in wb.sheetnames
    assert "05 Missing - bank" in wb.sheetnames
    assert "06 Field Differences" in wb.sheetnames
    assert "07 Control Checks" in wb.sheetnames
