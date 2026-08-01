from __future__ import annotations

import importlib.util
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def excel_writer_engine() -> str:
    return "xlsxwriter" if importlib.util.find_spec("xlsxwriter") else "openpyxl"


def auto_fit_columns(worksheet, df: pd.DataFrame) -> None:
    if df.empty:
        for i, col in enumerate(df.columns):
            width = min(len(str(col)) + 2, 60)
            if hasattr(worksheet, "set_column"):
                worksheet.set_column(i, i, width)
            else:
                worksheet.column_dimensions[get_column_letter(i + 1)].width = width
        return

    for i, col in enumerate(df.columns):
        col_series = df[col].fillna("").astype(str)
        max_len = max(int(col_series.str.len().max()), len(str(col))) + 2
        width = min(max_len, 60)
        if hasattr(worksheet, "set_column"):
            worksheet.set_column(i, i, width)
        else:
            worksheet.column_dimensions[get_column_letter(i + 1)].width = width


def style_header_row(workbook, worksheet, df: pd.DataFrame, color: str) -> None:
    if hasattr(workbook, "add_format"):
        header_fmt = workbook.add_format({"bold": True, "bg_color": color, "border": 1, "align": "center"})
        for col_num, col_name in enumerate(df.columns):
            worksheet.write(0, col_num, col_name, header_fmt)
        return

    fill = PatternFill("solid", fgColor=color.replace("#", ""))
    font = Font(bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_num, col_name in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=col_name)
        cell.font = font
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_generic_report(
    output_path: Path,
    reconciliation_results: list[dict],
    file_1_not_found: list[dict],
    file_2_not_found: list[dict],
    file_1_name: str = "File 1",
    file_2_name: str = "File 2",
) -> None:
    reconciliation_df = pd.DataFrame(reconciliation_results)
    file_1_not_found_df = pd.DataFrame(file_1_not_found)
    file_2_not_found_df = pd.DataFrame(file_2_not_found)

    sheet_f2 = _sheet_name(f"Not found in {file_2_name}")
    sheet_f1 = _sheet_name(f"Not found in {file_1_name}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine=excel_writer_engine()) as writer:
        reconciliation_df.to_excel(writer, sheet_name="Reconciliation Report", index=False)
        auto_fit_columns(writer.sheets["Reconciliation Report"], reconciliation_df)

        file_1_not_found_df.to_excel(writer, sheet_name=sheet_f2, index=False)
        auto_fit_columns(writer.sheets[sheet_f2], file_1_not_found_df)

        file_2_not_found_df.to_excel(writer, sheet_name=sheet_f1, index=False)
        auto_fit_columns(writer.sheets[sheet_f1], file_2_not_found_df)


def _sheet_name(label: str) -> str:
    return label[:31]


def _format_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy(deep=False)
    if "INVOICE DATE" in df.columns:
        df["INVOICE DATE"] = df["INVOICE DATE"].apply(
            lambda value: value.date().isoformat() if pd.notna(value) and hasattr(value, "date") else value
        )
    return df


def write_gst_output(
    mismatched: list,
    only_in_file1: list,
    only_in_file2: list,
    confidence_review: list,
    path: Path,
    file1_name: str = "File1",
    file2_name: str = "File2",
) -> None:
    mismatch_df = _format_date_columns(pd.DataFrame(mismatched))
    f1_df = _format_date_columns(pd.DataFrame(only_in_file1))
    f2_df = _format_date_columns(pd.DataFrame(only_in_file2))
    review_df = _format_date_columns(pd.DataFrame(confidence_review))

    sheet2_name = _sheet_name(f"{file1_name} Missing in {file2_name}")
    sheet3_name = _sheet_name(f"{file2_name} Missing in {file1_name}")

    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine=excel_writer_engine()) as writer:
        wb = writer.book
        sheets = [
            ("Mismatched Invoices", mismatch_df, "#FFC7CE"),
            (sheet2_name, f1_df, "#FFEB9C"),
            (sheet3_name, f2_df, "#C6EFCE"),
            ("Match Confidence Review", review_df, "#D9EAD3"),
        ]
        for sheet_name, df, color in sheets:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            style_header_row(wb, worksheet, df, color)
            auto_fit_columns(worksheet, df)


def generate_sample_format(path: Path, required_columns: list[str]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GST Data Template"

    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sample_data = [
        ["GSTR-2B", "ABC PRIVATE LIMITED", "INV-005", "2025-03-05", 10000, 1800, 0, 0, 0, 11800],
        ["GSTR-2B", "ABC PVT LTD", "INV005", "2025-03-05", 5000, 900, 0, 0, 0, 5900],
        ["GSTR-2B", "XYZ ENTERPRISES", "INV-042", "2025-03-10", 20000, 0, 1800, 1800, 0, 23600],
    ]
    col_widths = [14, 34, 22, 16, 16, 12, 12, 12, 12, 16]

    for col_idx, (col_name, width) in enumerate(zip(required_columns, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 35

    for row_idx, row_data in enumerate(sample_data, start=2):
        fill = PatternFill("solid", fgColor="DCE6F1" if row_idx % 2 == 0 else "FFFFFF")
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if col_idx <= 2 else "center", vertical="center")
            if col_idx == 4:
                cell.number_format = "DD-MMM-YYYY"
            elif col_idx >= 5:
                cell.number_format = "#,##0.00"

    wb.save(path)


def generate_generic_sample_format(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "General Data Template"

    columns = ["Invoice No", "Invoice Date", "Vendor Name", "Amount", "Department"]
    header_fill = PatternFill("solid", fgColor="087D72")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sample_data = [
        ["INV-1001", "2025-03-01", "ACME SUPPLIERS LTD", 15450.00, "OPERATIONS"],
        ["INV-1002", "2025-03-02", "GLOBAL TECH SOLUTIONS", 8200.50, "IT SERVICES"],
        ["INV-1003", "2025-03-05", "NEXUS LOGISTICS INC", 24300.00, "LOGISTICS"],
    ]
    col_widths = [18, 16, 30, 16, 18]

    for col_idx, (col_name, width) in enumerate(zip(columns, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 32

    for row_idx, row_data in enumerate(sample_data, start=2):
        fill = PatternFill("solid", fgColor="E4F4F1" if row_idx % 2 == 0 else "FFFFFF")
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if col_idx in (1, 3, 5) else "center", vertical="center")
            if col_idx == 2:
                cell.number_format = "DD-MMM-YYYY"
            elif col_idx == 4:
                cell.number_format = "#,##0.00"

    wb.save(path)
