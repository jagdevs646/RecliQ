from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, DoughnutChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def excel_writer_engine() -> str:
    return "openpyxl"


class ReportConfig:
    def __init__(
        self,
        include_summary: bool = True,
        include_exceptions: bool = True,
        include_matched: bool = True,
        include_missing_file_1: bool = True,
        include_missing_file_2: bool = True,
        include_field_differences: bool = True,
        include_controls: bool = True,
        date_format: str = "YYYY-MM-DD",
        number_format: str = "#,##0.00",
    ):
        self.include_summary = include_summary
        self.include_exceptions = include_exceptions
        self.include_matched = include_matched
        self.include_missing_file_1 = include_missing_file_1
        self.include_missing_file_2 = include_missing_file_2
        self.include_field_differences = include_field_differences
        self.include_controls = include_controls
        self.date_format = date_format
        self.number_format = number_format


def _clean_table_name(name: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_]", "", name)
    if cleaned and cleaned[0].isdigit():
        cleaned = f"_{cleaned}"
    return cleaned[:30] or "TableData"


def _clean_file_label(name: str) -> str:
    stem = Path(name).name
    if stem.lower().endswith(".xlsx"):
        return stem[:-5]
    if stem.lower().endswith(".xls"):
        return stem[:-4]
    return stem


class UniversalReporter:
    def __init__(self, data: dict, config: ReportConfig, output_path: Path):
        self.data = data
        self.config = config
        self.output_path = output_path
        self.font_family = "Arial"
        self.colors = {
            "primary": "1F3864",       # Deep Navy Blue
            "accent": "2E5395",        # Slate Blue
            "header_fg": "FFFFFF",     # White
            "pass": "1E7B34",          # Forest Green
            "warning": "B36A00",       # Amber / Orange
            "exception": "C0392B",     # Crimson Red
            "critical": "A6192E",      # Dark Crimson
            "portal_slice": "7A0C1E",  # Dark Maroon
            "field_tab": "B36A00",     # Amber / Orange
            "neutral_bg": "F2F2F2",    # Metric cards & zebra
            "text_dark": "000000",
            "text_muted": "595959",
            "border": "D9D9D9",
            "pass_bg": "D5F5E3",
            "fail_bg": "FADBD8",
        }

    def _thin_border(self, color: str | None = None) -> Border:
        c = color or self.colors["border"]
        s = Side(style="thin", color=c)
        return Border(left=s, right=s, top=s, bottom=s)

    def _sheet_name_missing(self, file_num: int) -> str:
        meta = self.data.get("metadata", {})
        fname = meta.get(f"file_{file_num}_name", f"Source {file_num}")
        clean_name = _clean_file_label(fname)
        num_prefix = "04" if file_num == 1 else "05"
        full = f"{num_prefix} Missing - {clean_name}"
        return full[:31]

    def generate(self) -> None:
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        # Track initial active sheet to remove later
        default_sheet = wb.active

        sheet_missing_1 = self._sheet_name_missing(1)
        sheet_missing_2 = self._sheet_name_missing(2)

        if self.config.include_summary:
            self._generate_summary(wb, sheet_missing_1, sheet_missing_2)
        if self.config.include_exceptions:
            self._generate_exceptions(wb)
        if self.config.include_matched:
            self._generate_matched(wb)
        if self.config.include_missing_file_1:
            self._generate_missing(wb, 1, sheet_missing_1)
        if self.config.include_missing_file_2:
            self._generate_missing(wb, 2, sheet_missing_2)
        if self.config.include_field_differences:
            self._generate_field_differences(wb)
        if self.config.include_controls:
            self._generate_controls(wb, sheet_missing_1, sheet_missing_2)

        if default_sheet and default_sheet in wb.worksheets:
            wb.remove(default_sheet)

        if len(wb.worksheets) == 0:
            ws = wb.create_sheet("Empty")
            ws.cell(row=1, column=1, value="No sections selected")

        # Set active sheet to summary if present
        if "01 Executive Summary" in wb.sheetnames:
            wb.active = wb["01 Executive Summary"]

        wb.save(self.output_path)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1: 01 Executive Summary
    # ═══════════════════════════════════════════════════════════════════════
    def _generate_summary(self, wb: openpyxl.Workbook, sheet_missing_1: str, sheet_missing_2: str) -> None:  # noqa: C901
        ws = wb.create_sheet("01 Executive Summary")
        ws.sheet_properties.tabColor = self.colors["primary"]
        ws.views.sheetView[0].showGridLines = False

        meta = self.data.get("metadata", {})
        file1_name = meta.get("file_1_name", "File 1")
        file2_name = meta.get("file_2_name", "File 2")
        matching_keys = meta.get("matching_keys", ["Key"])
        key_label = ", ".join(matching_keys) if matching_keys else "Key"

        # Format generated date
        gen_time_str = datetime.now(timezone.utc).strftime("%B %d, %Y")

        # Row counts in detail tabs
        exceptions_list = self.data.get("exceptions", [])
        matched_list = self.data.get("matched_records", [])
        missing_1_list = self.data.get("missing_in_file_1", [])
        missing_2_list = self.data.get("missing_in_file_2", [])
        field_exc = self.data.get("field_exception_summary", [])

        max_exc_row = max(5, len(exceptions_list) + 4)
        max_matched_row = max(5, len(matched_list) + 4)
        max_m1_row = max(5, len(missing_1_list) + 4)
        max_m2_row = max(5, len(missing_2_list) + 4)

        border = self._thin_border()
        navy_fill = PatternFill("solid", fgColor=self.colors["primary"])
        slate_fill = PatternFill("solid", fgColor=self.colors["accent"])
        card_fill = PatternFill("solid", fgColor=self.colors["neutral_bg"])

        # ── Row 1: Main Header Banner ──────────────────────────────────────
        ws.merge_cells("B1:N1")
        b1 = ws["B1"]
        b1.value = "RECLIQ  |  RECONCILIATION EXECUTIVE SUMMARY"
        b1.font = Font(name=self.font_family, size=18, bold=True, color=self.colors["header_fg"])
        b1.fill = navy_fill
        b1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 33.75

        # ── Row 2: Subtitle Metadata ───────────────────────────────────────
        ws.merge_cells("B2:N2")
        b2 = ws["B2"]
        b2.value = f"{file1_name}  vs.  {file2_name}     •     Matching Key: {key_label}     •     Report Generated: {gen_time_str}"
        b2.font = Font(name=self.font_family, size=10, color=self.colors["header_fg"])
        b2.fill = slate_fill
        b2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 19.5

        # ── Row 4: KEY METRICS Section ─────────────────────────────────────
        ws.merge_cells("B4:N4")
        b4 = ws["B4"]
        b4.value = "KEY METRICS"
        b4.font = Font(name=self.font_family, size=12, bold=True, color=self.colors["primary"])
        b4.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[4].height = 21.75

        # ── Row 5: KPI Headers ─────────────────────────────────────────────
        kpi_headers = [
            ("B5:C5", "TOTAL RECORDS IN SCOPE"),
            ("D5:E5", "MATCHED"),
            ("F5:G5", "EXCEPTIONS"),
            ("H5:I5", "MISSING"),
            ("J5:K5", "MATCH RATE"),
            ("L5:M5", "EXCEPTION RATE"),
        ]
        for rng, label in kpi_headers:
            ws.merge_cells(rng)
            c = ws[rng.split(":")[0]]
            c.value = label
            c.font = Font(name=self.font_family, size=8, bold=True, color=self.colors["text_muted"])
            c.fill = card_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[5].height = 15.75

        # ── Row 6: KPI Values with dynamic formulas ────────────────────────
        kpi_values = [
            (
                "B6:C6",
                f"=COUNTA('03 Matched Records'!C5:C{max_matched_row})+SUM('02 Exceptions'!L5:L{max_exc_row})"
                f"+COUNTA('{sheet_missing_1}'!A5:A{max_m1_row})+COUNTA('{sheet_missing_2}'!A5:A{max_m2_row})",
                self.colors["primary"],
                "#,##0",
            ),
            ("D6:E6", f"=COUNTA('03 Matched Records'!C5:C{max_matched_row})", self.colors["pass"], "#,##0"),
            ("F6:G6", f"=SUM('02 Exceptions'!L5:L{max_exc_row})", self.colors["warning"], "#,##0"),
            (
                "H6:I6",
                f"=COUNTA('{sheet_missing_1}'!A5:A{max_m1_row})+COUNTA('{sheet_missing_2}'!A5:A{max_m2_row})",
                self.colors["critical"],
                "#,##0",
            ),
            ("J6:K6", "=IF($B$6>0,$D$6/$B$6,0)", self.colors["pass"], "0.0%"),
            ("L6:M6", "=IF($B$6>0,$F$6/$B$6,0)", self.colors["warning"], "0.0%"),
        ]
        for rng, form, color, num_fmt in kpi_values:
            ws.merge_cells(rng)
            c = ws[rng.split(":")[0]]
            c.value = form
            c.font = Font(name=self.font_family, size=20, bold=True, color=color)
            c.fill = card_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = num_fmt
        ws.row_dimensions[6].height = 30

        # Apply borders around KPI cards
        for col_idx in range(2, 14):
            for row_idx in (5, 6, 7):
                ws.cell(row=row_idx, column=col_idx).border = border

        # ── Row 7: KPI Subtext / dynamic text formulas ─────────────────────
        kpi_subtext = [
            ("B7:C7", f'="Unique {key_label} across both files"'),
            ("D7:E7", '="of "&$B$6&" records reviewed"'),
            ("F7:G7", f'=(COUNTA(\'02 Exceptions\'!B5:B{max_exc_row}))&" field-level mismatch(es)"'),
            ("H7:I7", f'=(COUNTA(\'{sheet_missing_1}\'!A5:A{max_m1_row}))&" + "&(COUNTA(\'{sheet_missing_2}\'!A5:A{max_m2_row}))&" by source"'),
            ("J7:K7", '="records matched cleanly"'),
            ("L7:M7", '="records need review"'),
        ]
        for rng, form in kpi_subtext:
            ws.merge_cells(rng)
            c = ws[rng.split(":")[0]]
            c.value = form
            c.font = Font(name=self.font_family, size=8, color=self.colors["text_muted"])
            c.fill = card_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[7].height = 15.75

        # ── Row 9: RECONCILIATION OUTCOME ──────────────────────────────────
        ws.merge_cells("B9:F9")
        b9 = ws["B9"]
        b9.value = "RECONCILIATION OUTCOME"
        b9.font = Font(name=self.font_family, size=12, bold=True, color=self.colors["primary"])
        b9.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[9].height = 21.75

        # Outcome Table Header (Row 10)
        ws.row_dimensions[10].height = 15.75
        for col_idx, h_text in enumerate(["Category", "Count", "% of Total"], start=2):
            cell = ws.cell(row=10, column=col_idx, value=h_text)
            cell.font = Font(name=self.font_family, size=10, bold=True, color=self.colors["header_fg"])
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Outcome Data Rows (Rows 11-14)
        outcome_rows = [
            ("Matched", "=$D$6"),
            ("Exceptions", "=$F$6"),
            (f"Missing — {file1_name}", f"=COUNTA('{sheet_missing_1}'!A5:A{max_m1_row})"),
            (f"Missing — {file2_name}", f"=COUNTA('{sheet_missing_2}'!A5:A{max_m2_row})"),
        ]
        for idx, (cat_name, count_form) in enumerate(outcome_rows, start=11):
            fill = card_fill if idx % 2 == 0 else PatternFill(fill_type=None)
            c_cat = ws.cell(row=idx, column=2, value=cat_name)
            c_cat.font = Font(name=self.font_family, size=9)
            c_cat.fill = fill
            c_cat.alignment = Alignment(horizontal="left", vertical="center")
            c_cat.border = border

            c_count = ws.cell(row=idx, column=3, value=count_form)
            c_count.font = Font(name=self.font_family, size=9)
            c_count.fill = fill
            c_count.alignment = Alignment(horizontal="center", vertical="center")
            c_count.number_format = "#,##0"
            c_count.border = border

            c_pct = ws.cell(row=idx, column=4, value=f"=C{idx}/$B$6")
            c_pct.font = Font(name=self.font_family, size=9)
            c_pct.fill = fill
            c_pct.alignment = Alignment(horizontal="center", vertical="center")
            c_pct.number_format = "0.0%"
            c_pct.border = border

        # Outcome Total Row (Row 15)
        ws.cell(row=15, column=2, value="Total").font = Font(name=self.font_family, size=9, bold=True)
        ws.cell(row=15, column=2).border = border

        c15 = ws.cell(row=15, column=3, value="=SUM(C11:C14)")
        c15.font = Font(name=self.font_family, size=9, bold=True)
        c15.alignment = Alignment(horizontal="center", vertical="center")
        c15.number_format = "#,##0"
        c15.border = border

        d15 = ws.cell(row=15, column=4, value="=SUM(D11:D14)")
        d15.font = Font(name=self.font_family, size=9, bold=True)
        d15.alignment = Alignment(horizontal="center", vertical="center")
        d15.number_format = "0.0%"
        d15.border = border

        # ── Outcome Doughnut / Pie Chart (Position: G8 to M22) ─────────────
        pie = DoughnutChart()
        pie.title = "Reconciliation Outcome"
        pie.holeSize = 50
        labels_ref = Reference(ws, min_col=2, min_row=11, max_row=14)
        data_ref = Reference(ws, min_col=3, min_row=10, max_row=14)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels_ref)
        pie.legend.legendPos = "b"
        pie.width = 17
        pie.height = 10.5

        pie_colors = [self.colors["pass"], self.colors["warning"], self.colors["critical"], self.colors["portal_slice"]]
        for i, color in enumerate(pie_colors):
            dp = DataPoint(idx=i)
            dp.graphicalProperties.solidFill = color
            pie.series[0].data_points.append(dp)

        ws.add_chart(pie, "G8")

        # ── Row 17: EXCEPTION BREAKDOWN BY TYPE ────────────────────────────
        ws.merge_cells("B17:F17")
        b17 = ws["B17"]
        b17.value = "EXCEPTION BREAKDOWN BY TYPE"
        b17.font = Font(name=self.font_family, size=12, bold=True, color=self.colors["primary"])
        b17.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[17].height = 21.75

        # Table Header (Row 18)
        ws.row_dimensions[18].height = 15.75
        for col_idx, h_text in enumerate(["Exception Type", "Count", "% of Total"], start=2):
            cell = ws.cell(row=18, column=col_idx, value=h_text)
            cell.font = Font(name=self.font_family, size=10, bold=True, color=self.colors["header_fg"])
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Breakdown Data Rows (Rows 19-22)
        exc_breakdown = [
            ("Value Difference", f"=COUNTIFS('02 Exceptions'!$H$5:$H${max_exc_row},\"Value Difference\")"),
            ("Text Difference", f"=COUNTIFS('02 Exceptions'!$H$5:$H${max_exc_row},\"Text Difference\")"),
            (f"Missing — {file1_name}", f"=COUNTA('{sheet_missing_1}'!A5:A{max_m1_row})"),
            (f"Missing — {file2_name}", f"=COUNTA('{sheet_missing_2}'!A5:A{max_m2_row})"),
        ]
        for idx, (exc_type, count_form) in enumerate(exc_breakdown, start=19):
            fill = card_fill if idx % 2 == 0 else PatternFill(fill_type=None)
            c_type = ws.cell(row=idx, column=2, value=exc_type)
            c_type.font = Font(name=self.font_family, size=9)
            c_type.fill = fill
            c_type.alignment = Alignment(horizontal="left", vertical="center")
            c_type.border = border

            c_count = ws.cell(row=idx, column=3, value=count_form)
            c_count.font = Font(name=self.font_family, size=9)
            c_count.fill = fill
            c_count.alignment = Alignment(horizontal="center", vertical="center")
            c_count.number_format = "#,##0"
            c_count.border = border

            c_pct = ws.cell(row=idx, column=4, value=f"=C{idx}/SUM($C$19:$C$22)")
            c_pct.font = Font(name=self.font_family, size=9)
            c_pct.fill = fill
            c_pct.alignment = Alignment(horizontal="center", vertical="center")
            c_pct.number_format = "0.0%"
            c_pct.border = border

        # ── Row 24: FIELD-LEVEL PERFORMANCE ────────────────────────────────
        ws.merge_cells("B24:F24")
        b24 = ws["B24"]
        b24.value = "FIELD-LEVEL PERFORMANCE"
        b24.font = Font(name=self.font_family, size=12, bold=True, color=self.colors["primary"])
        b24.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[24].height = 21.75

        # Table Header (Row 25)
        ws.row_dimensions[25].height = 15.75
        for col_idx, h_text in enumerate(["Field", "Matched", "Mismatches", "Match Rate"], start=2):
            cell = ws.cell(row=25, column=col_idx, value=h_text)
            cell.font = Font(name=self.font_family, size=10, bold=True, color=self.colors["header_fg"])
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        fields = [f.get("Field", "") for f in field_exc if f.get("Field")] or ["Sample Field"]
        fld_start_row = 26
        fld_end_row = fld_start_row + len(fields) - 1

        for idx, f_name in enumerate(fields, start=fld_start_row):
            fill = card_fill if idx % 2 == 1 else PatternFill(fill_type=None)
            c_f = ws.cell(row=idx, column=2, value=f_name)
            c_f.font = Font(name=self.font_family, size=9)
            c_f.fill = fill
            c_f.alignment = Alignment(horizontal="left", vertical="center")
            c_f.border = border

            c_m = ws.cell(row=idx, column=3, value="=$D$6")
            c_m.font = Font(name=self.font_family, size=9)
            c_m.fill = fill
            c_m.alignment = Alignment(horizontal="center", vertical="center")
            c_m.number_format = "#,##0"
            c_m.border = border

            c_mm = ws.cell(row=idx, column=4, value=f"=COUNTIFS('02 Exceptions'!$C$5:$C${max_exc_row},B{idx})")
            c_mm.font = Font(name=self.font_family, size=9)
            c_mm.fill = fill
            c_mm.alignment = Alignment(horizontal="center", vertical="center")
            c_mm.number_format = "#,##0"
            c_mm.border = border

            c_rate = ws.cell(row=idx, column=5, value=f"=IF((C{idx}+D{idx})>0,C{idx}/(C{idx}+D{idx}),0)")
            c_rate.font = Font(name=self.font_family, size=9)
            c_rate.fill = fill
            c_rate.alignment = Alignment(horizontal="center", vertical="center")
            c_rate.number_format = "0.0%"
            c_rate.border = border

        # ── Column Bar Chart: Field-Level Mismatches (Position: G23 to M36) ─
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Field-Level Mismatches"
        bar.y_axis.title = None
        bar.x_axis.title = None
        bar.legend = None
        bar.width = 17
        bar.height = 10.5

        bar_cats = Reference(ws, min_col=2, min_row=fld_start_row, max_row=fld_end_row)
        bar_data = Reference(ws, min_col=4, min_row=25, max_row=fld_end_row)
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(bar_cats)
        if bar.series:
            bar.series[0].graphicalProperties.solidFill = self.colors["primary"]

        ws.add_chart(bar, "G23")

        # ── Row 30: KEY INSIGHTS ───────────────────────────────────────────
        ins_start_row = max(30, fld_end_row + 2)
        ws.merge_cells(start_row=ins_start_row, start_column=2, end_row=ins_start_row, end_column=13)
        b_ins = ws.cell(row=ins_start_row, column=2, value="KEY INSIGHTS")
        b_ins.font = Font(name=self.font_family, size=12, bold=True, color=self.colors["primary"])
        b_ins.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[ins_start_row].height = 21.75

        # Dynamic Formula Bullets
        bullets = [
            '="•  "&TEXT($J$6,"0.0%")&" of records matched perfectly ("&$D$6&" of "&$B$6&")."',
            f'="•  "&$F$6&" discrepant record(s) identified ("&(COUNTA(\'02 Exceptions\'!B5:B{max_exc_row}))&" field mismatches) — review required."',
            f'="•  "&(COUNTA(\'{sheet_missing_1}\'!A5:A{max_m1_row})+COUNTA(\'{sheet_missing_2}\'!A5:A{max_m2_row}))&" record(s) missing from one or both sources ("&(COUNTA(\'{sheet_missing_1}\'!A5:A{max_m1_row}))&" {file1_name}, "&(COUNTA(\'{sheet_missing_2}\'!A5:A{max_m2_row}))&" {file2_name})."',
            f'="•  Field \'"&INDEX($B${fld_start_row}:$B${fld_end_row},MATCH(MAX($D${fld_start_row}:$D${fld_end_row}),$D${fld_start_row}:$D${fld_end_row},0))&"\' has the highest exception count ("&MAX($D${fld_start_row}:$D${fld_end_row})&")."',
        ]
        for i, b_form in enumerate(bullets, start=ins_start_row + 1):
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=14)
            c = ws.cell(row=i, column=2, value=b_form)
            c.font = Font(name=self.font_family, size=9.5)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[i].height = 15.75

        # ── CONTROL SUMMARY Section ────────────────────────────────────────
        ctrl_start = ins_start_row + 5 + 1
        ws.merge_cells(start_row=ctrl_start, start_column=2, end_row=ctrl_start, end_column=13)
        b_ctrl = ws.cell(row=ctrl_start, column=2, value="CONTROL SUMMARY")
        b_ctrl.font = Font(name=self.font_family, size=12, bold=True, color=self.colors["primary"])
        b_ctrl.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[ctrl_start].height = 21.75

        ctrl_headers = [
            (f"B{ctrl_start+1}:D{ctrl_start+1}", "Total Records"),
            (f"E{ctrl_start+1}:G{ctrl_start+1}", "Outstanding Exceptions"),
            (f"H{ctrl_start+1}:J{ctrl_start+1}", "Missing Records"),
            (f"K{ctrl_start+1}:M{ctrl_start+1}", "Full Audit Trail"),
        ]
        for rng, lbl in ctrl_headers:
            ws.merge_cells(rng)
            c = ws[rng.split(":")[0]]
            c.value = lbl
            c.font = Font(name=self.font_family, size=8, bold=True, color=self.colors["text_muted"])
            c.fill = card_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ctrl_start + 1].height = 15.75

        ctrl_values = [
            (f"B{ctrl_start+2}:D{ctrl_start+2}", "=$B$6", "#,##0"),
            (f"E{ctrl_start+2}:G{ctrl_start+2}", "=$F$6", "#,##0"),
            (f"H{ctrl_start+2}:J{ctrl_start+2}", "=$H$6", "#,##0"),
            (f"K{ctrl_start+2}:M{ctrl_start+2}", "See tab 07", None),
        ]
        for rng, val, num_fmt in ctrl_values:
            ws.merge_cells(rng)
            c = ws[rng.split(":")[0]]
            c.value = val
            c.font = Font(name=self.font_family, size=13, bold=True, color=self.colors["primary"])
            c.fill = card_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            if num_fmt:
                c.number_format = num_fmt
        ws.row_dimensions[ctrl_start + 2].height = 21.75

        for col_idx in range(2, 14):
            for row_idx in (ctrl_start + 1, ctrl_start + 2):
                ws.cell(row=row_idx, column=col_idx).border = border

        # Footers
        f_row1 = ctrl_start + 4
        ws.merge_cells(start_row=f_row1, start_column=2, end_row=f_row1, end_column=13)
        c_f1 = ws.cell(
            row=f_row1,
            column=2,
            value="Full detail, formulas, and audit trail for every figure above are available on tabs 02–07 of this workbook.",
        )
        c_f1.font = Font(name=self.font_family, size=8, color=self.colors["text_muted"])

        f_row2 = f_row1 + 2
        ws.merge_cells(start_row=f_row2, start_column=2, end_row=f_row2, end_column=13)
        c_f2 = ws.cell(
            row=f_row2,
            column=2,
            value="RecliQ Reconciliation Engine  •  Confidential — Internal Use Only",
        )
        c_f2.font = Font(name=self.font_family, size=8, color=self.colors["text_muted"])
        c_f2.alignment = Alignment(horizontal="center", vertical="center")

        # Column widths
        ws.column_dimensions["A"].width = 2.44
        ws.column_dimensions["B"].width = 22.0
        for col_ltr in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
            ws.column_dimensions[col_ltr].width = 13.0
        ws.column_dimensions["N"].width = 2.44

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2: 02 Exceptions
    # ═══════════════════════════════════════════════════════════════════════
    def _generate_exceptions(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("02 Exceptions")
        ws.sheet_properties.tabColor = self.colors["exception"]
        ws.views.sheetView[0].showGridLines = False

        meta = self.data.get("metadata", {})
        file1_name = meta.get("file_1_name", "File 1")
        file2_name = meta.get("file_2_name", "File 2")
        matching_keys = meta.get("matching_keys", ["Key"])
        key_label = ", ".join(matching_keys) if matching_keys else "Key"

        # Row 1: Banner Title
        ws.merge_cells("A1:L1")
        a1 = ws["A1"]
        a1.value = "Exception Log — Field-Level Mismatches"
        a1.font = Font(name=self.font_family, size=13, bold=True, color=self.colors["primary"])
        ws.row_dimensions[1].height = 19.5

        # Row 2: Subtitle
        ws.merge_cells("A2:L2")
        a2 = ws["A2"]
        a2.value = "Records present in both files where one or more compared fields did not match. One row per field-level mismatch; a record may appear more than once."
        a2.font = Font(name=self.font_family, size=9, color=self.colors["text_muted"])
        ws.row_dimensions[2].height = 15.75

        # Row 4: Header
        headers = [
            "Exception ID",
            f"Match Key ({key_label})",
            "Field",
            f"{file1_name} Value",
            f"{file2_name} Value",
            "Difference",
            "Difference %",
            "Exception Type",
            "Severity",
            "Status",
            "Action Notes",
            "Distinct Record",
        ]
        ws.row_dimensions[4].height = 19.5
        navy_fill = PatternFill("solid", fgColor=self.colors["primary"])
        border = self._thin_border()

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = Font(name=self.font_family, size=10, bold=True, color=self.colors["header_fg"])
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        exceptions_list = self.data.get("exceptions", [])
        zebra_fill = PatternFill("solid", fgColor=self.colors["neutral_bg"])

        for row_idx, exc in enumerate(exceptions_list, start=5):
            fill = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
            diff_val = exc.get("Difference")
            diff_pct = exc.get("Difference %")

            # Convert diff_pct to decimal if numeric percentage
            diff_pct_val = None
            if diff_pct is not None:
                try:
                    diff_pct_val = float(str(diff_pct).replace("%", "")) / 100.0 if float(str(diff_pct).replace("%", "")) > 1.0 or "%" in str(diff_pct) else float(diff_pct)
                except Exception:
                    diff_pct_val = None

            row_data = [
                (exc.get("Exception ID", f"EX-{str(row_idx-4).zfill(6)}"), "left", None),
                (str(exc.get("Match Key", "")), "left", None),
                (str(exc.get("Field", "")), "left", None),
                (exc.get("File 1 Value"), "right" if isinstance(exc.get("File 1 Value"), (int, float)) else "left", self.config.number_format if isinstance(exc.get("File 1 Value"), (int, float)) else None),
                (exc.get("File 2 Value"), "right" if isinstance(exc.get("File 2 Value"), (int, float)) else "left", self.config.number_format if isinstance(exc.get("File 2 Value"), (int, float)) else None),
                (diff_val, "right", self.config.number_format if diff_val is not None else None),
                (diff_pct_val, "right", "0.0%" if diff_pct_val is not None else None),
                (exc.get("Exception Type", "Value Difference"), "left", None),
                (exc.get("Severity", "Medium"), "left", None),
                (exc.get("Status", "Open"), "left", None),
                (exc.get("Action Notes", ""), "left", None),
                (f"=IF(COUNTIF($B$5:B{row_idx},B{row_idx})=1,1,0)", "right", None),
            ]

            for col_idx, (val, align_h, num_fmt) in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name=self.font_family, size=9)
                cell.fill = fill
                cell.alignment = Alignment(horizontal=align_h, vertical="center")
                cell.border = border
                if num_fmt:
                    cell.number_format = num_fmt

        last_row = max(4, len(exceptions_list) + 4)
        ws.freeze_panes = "A5"
        if len(exceptions_list) > 0:
            ws.auto_filter.ref = f"A4:K{last_row}"

        col_widths = [16.5, 21.5, 16.0, 20.5, 20.0, 14.5, 16.5, 19.0, 12.5, 11.5, 16.5, 14.5]
        for col_idx, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 3: 03 Matched Records
    # ═══════════════════════════════════════════════════════════════════════
    def _generate_matched(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("03 Matched Records")
        ws.sheet_properties.tabColor = self.colors["pass"]
        ws.views.sheetView[0].showGridLines = False

        meta = self.data.get("metadata", {})
        file1_name = meta.get("file_1_name", "File 1")
        file2_name = meta.get("file_2_name", "File 2")
        matching_keys = meta.get("matching_keys", ["Key"])
        key_label = ", ".join(matching_keys) if matching_keys else "Key"

        # Row 1: Banner Title
        ws.merge_cells("A1:G1")
        a1 = ws["A1"]
        a1.value = "Matched Records — Audit Log"
        a1.font = Font(name=self.font_family, size=13, bold=True, color=self.colors["primary"])
        ws.row_dimensions[1].height = 19.5

        # Row 2: Subtitle
        ws.merge_cells("A2:G2")
        a2 = ws["A2"]
        a2.value = "Records that matched perfectly across all compared fields. Read-only audit log."
        a2.font = Font(name=self.font_family, size=9, color=self.colors["text_muted"])
        ws.row_dimensions[2].height = 15.75

        # Row 4: Header
        headers = [
            f"{file1_name} Row",
            f"{file2_name} Row",
            f"{key_label} ({file1_name})",
            f"{key_label} ({file2_name})",
            "Match Type",
            "Match Confidence",
            "Match Status",
        ]
        ws.row_dimensions[4].height = 19.5
        navy_fill = PatternFill("solid", fgColor=self.colors["primary"])
        border = self._thin_border()

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = Font(name=self.font_family, size=10, bold=True, color=self.colors["header_fg"])
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        matched_list = self.data.get("matched_records", [])
        zebra_fill = PatternFill("solid", fgColor=self.colors["neutral_bg"])

        for row_idx, row in enumerate(matched_list, start=5):
            fill = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
            r1 = row.get("ROW (FILE 1)", row.get("ROW (File 1)", row_idx - 3))
            r2 = row.get("ROW (FILE 2)", row.get("ROW (File 2)", row_idx - 3))
            k1 = row.get(key_label, row.get(f"{key_label} (File1)", row.get(f"{key_label} (FILE 1)", "")))
            k2 = row.get(f"MATCHED {key_label}", row.get(f"{key_label} (File2)", row.get(f"{key_label} (FILE 2)", k1)))

            row_data = [
                (r1, "center"),
                (r2, "center"),
                (str(k1), "left"),
                (str(k2), "left"),
                (row.get("MATCH TYPE", "exact"), "center"),
                (str(row.get("MATCH CONFIDENCE", "100%")), "center"),
                (row.get("MATCH STATUS", "Exact Match"), "center"),
            ]

            for col_idx, (val, align_h) in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name=self.font_family, size=9)
                cell.fill = fill
                cell.alignment = Alignment(horizontal=align_h, vertical="center")
                cell.border = border

        last_row = max(4, len(matched_list) + 4)
        ws.freeze_panes = "A5"

        if len(matched_list) > 0:
            tab = Table(displayName="MatchedRecordsTbl", ref=f"A4:G{last_row}")
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            ws.add_table(tab)

        col_widths = [16.0, 16.0, 22.0, 22.0, 14.0, 16.0, 16.0]
        for col_idx, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 4 & 5: 04 Missing - File 1 & 05 Missing - File 2
    # ═══════════════════════════════════════════════════════════════════════
    def _generate_missing(self, wb: openpyxl.Workbook, file_num: int, sheet_name: str) -> None:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = self.colors["critical"]
        ws.views.sheetView[0].showGridLines = False

        meta = self.data.get("metadata", {})
        f1_name = meta.get("file_1_name", "File 1")
        f2_name = meta.get("file_2_name", "File 2")
        this_file = f1_name if file_num == 1 else f2_name
        other_file = f2_name if file_num == 1 else f1_name
        matching_keys = meta.get("matching_keys", ["Key"])
        key_label = ", ".join(matching_keys) if matching_keys else "Key"

        # Data key: items missing in File 1 come from File 2
        key = "missing_in_file_1" if file_num == 1 else "missing_in_file_2"
        missing_rows = self.data.get(key, [])

        # Determine all column names from missing rows
        cols = []
        if missing_rows:
            raw_cols = list(missing_rows[0].keys())
            # Ensure ROW column is at the end
            cols = [c for c in raw_cols if not c.startswith("ROW")]
            row_col = next((c for c in raw_cols if c.startswith("ROW")), None)
            if row_col:
                cols.append(f"{other_file} Row")
        else:
            cols = [key_label, f"{other_file} Row"]

        last_col_ltr = get_column_letter(max(len(cols), 1))

        # Row 1: Banner Title
        ws.merge_cells(f"A1:{last_col_ltr}1")
        a1 = ws["A1"]
        a1.value = f"Missing from {this_file} (File {file_num})"
        a1.font = Font(name=self.font_family, size=13, bold=True, color=self.colors["primary"])
        ws.row_dimensions[1].height = 19.5

        # Row 2: Subtitle
        ws.merge_cells(f"A2:{last_col_ltr}2")
        a2 = ws["A2"]
        a2.value = f"Records present in {other_file} but with no matching {key_label} found in {this_file}."
        a2.font = Font(name=self.font_family, size=9, color=self.colors["text_muted"])
        ws.row_dimensions[2].height = 15.75

        # Row 4: Header
        ws.row_dimensions[4].height = 27.75
        navy_fill = PatternFill("solid", fgColor=self.colors["primary"])
        border = self._thin_border()

        for col_idx, h in enumerate(cols, start=1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = Font(name=self.font_family, size=10, bold=True, color=self.colors["header_fg"])
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        zebra_fill = PatternFill("solid", fgColor=self.colors["neutral_bg"])

        for row_idx, row_dict in enumerate(missing_rows, start=5):
            fill = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
            for col_idx, col_name in enumerate(cols, start=1):
                raw_key = next((k for k in row_dict if k == col_name or (col_name.endswith("Row") and k.startswith("ROW"))), col_name)
                val = row_dict.get(raw_key)

                is_num = isinstance(val, (int, float)) and not isinstance(val, bool)
                align_h = "right" if is_num else "left"

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name=self.font_family, size=9)
                cell.fill = fill
                cell.alignment = Alignment(horizontal=align_h, vertical="center")
                cell.border = border
                if is_num and not col_name.endswith("Row"):
                    cell.number_format = self.config.number_format

        last_row = max(4, len(missing_rows) + 4)
        ws.freeze_panes = "A5"

        if len(missing_rows) > 0:
            tbl_name = _clean_table_name(f"Missing_{sheet_name}")
            tab = Table(displayName=tbl_name, ref=f"A4:{last_col_ltr}{last_row}")
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            ws.add_table(tab)

        for col_idx, col_name in enumerate(cols, start=1):
            w = max(len(str(col_name)) + 3, 12)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(w, 35)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 6: 06 Field Differences
    # ═══════════════════════════════════════════════════════════════════════
    def _generate_field_differences(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("06 Field Differences")
        ws.sheet_properties.tabColor = self.colors["field_tab"]
        ws.views.sheetView[0].showGridLines = False

        meta = self.data.get("metadata", {})
        file1_name = meta.get("file_1_name", "File 1")
        file2_name = meta.get("file_2_name", "File 2")
        matching_keys = meta.get("matching_keys", ["Key"])
        key_label = ", ".join(matching_keys) if matching_keys else "Key"

        # Row 1: Banner Title
        ws.merge_cells("A1:G1")
        a1 = ws["A1"]
        a1.value = "Field Differences — Granular Side-by-Side Comparison"
        a1.font = Font(name=self.font_family, size=13, bold=True, color=self.colors["primary"])
        ws.row_dimensions[1].height = 19.5

        # Row 2: Subtitle
        ws.merge_cells("A2:G2")
        a2 = ws["A2"]
        a2.value = "Detailed breakdown of every field-level difference between matched record pairs."
        a2.font = Font(name=self.font_family, size=9, color=self.colors["text_muted"])
        ws.row_dimensions[2].height = 15.75

        # Row 4: Header
        headers = [
            f"Match Key ({key_label})",
            "Field",
            f"{file1_name} Value",
            f"{file2_name} Value",
            "Difference",
            "Difference %",
            "Result",
        ]
        ws.row_dimensions[4].height = 19.5
        navy_fill = PatternFill("solid", fgColor=self.colors["primary"])
        border = self._thin_border()

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = Font(name=self.font_family, size=10, bold=True, color=self.colors["header_fg"])
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        fld_diffs = self.data.get("field_differences", [])
        zebra_fill = PatternFill("solid", fgColor=self.colors["neutral_bg"])

        for row_idx, row in enumerate(fld_diffs, start=5):
            fill = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
            diff_val = row.get("Difference")
            diff_pct = row.get("Difference %")

            diff_pct_val = None
            if diff_pct is not None:
                try:
                    diff_pct_val = float(str(diff_pct).replace("%", "")) / 100.0 if float(str(diff_pct).replace("%", "")) > 1.0 or "%" in str(diff_pct) else float(diff_pct)
                except Exception:
                    diff_pct_val = None

            row_data = [
                (str(row.get("Match Key", "")), "left", None),
                (str(row.get("Field", "")), "left", None),
                (row.get("File 1 Value"), "right" if isinstance(row.get("File 1 Value"), (int, float)) else "left", self.config.number_format if isinstance(row.get("File 1 Value"), (int, float)) else None),
                (row.get("File 2 Value"), "right" if isinstance(row.get("File 2 Value"), (int, float)) else "left", self.config.number_format if isinstance(row.get("File 2 Value"), (int, float)) else None),
                (diff_val, "right", self.config.number_format if diff_val is not None else None),
                (diff_pct_val, "right", "0.0%" if diff_pct_val is not None else None),
                (str(row.get("Result", "Mismatch")), "left", None),
            ]

            for col_idx, (val, align_h, num_fmt) in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name=self.font_family, size=9)
                cell.fill = fill
                cell.alignment = Alignment(horizontal=align_h, vertical="center")
                cell.border = border
                if num_fmt:
                    cell.number_format = num_fmt

        ws.freeze_panes = "A5"
        col_widths = [18.0, 18.0, 18.0, 18.0, 14.0, 14.0, 14.0]
        for col_idx, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 7: 07 Control Checks
    # ═══════════════════════════════════════════════════════════════════════
    def _generate_controls(self, wb: openpyxl.Workbook, sheet_missing_1: str, sheet_missing_2: str) -> None:
        ws = wb.create_sheet("07 Control Checks")
        ws.sheet_properties.tabColor = self.colors["primary"]
        ws.views.sheetView[0].showGridLines = False

        meta = self.data.get("metadata", {})
        file1_name = meta.get("file_1_name", "File 1")
        file2_name = meta.get("file_2_name", "File 2")
        matching_keys = meta.get("matching_keys", ["Key"])
        key_label = ", ".join(matching_keys) if matching_keys else "Key"

        exceptions_list = self.data.get("exceptions", [])
        matched_list = self.data.get("matched_records", [])
        missing_1_list = self.data.get("missing_in_file_1", [])
        missing_2_list = self.data.get("missing_in_file_2", [])

        max_exc_row = max(5, len(exceptions_list) + 4)
        max_matched_row = max(5, len(matched_list) + 4)
        max_m1_row = max(5, len(missing_1_list) + 4)
        max_m2_row = max(5, len(missing_2_list) + 4)

        # Row 1: Banner Title
        ws.merge_cells("A1:E1")
        a1 = ws["A1"]
        a1.value = "Control Checks — Reconciliation Audit Trail"
        a1.font = Font(name=self.font_family, size=13, bold=True, color=self.colors["primary"])
        ws.row_dimensions[1].height = 19.5

        # Row 2: Subtitle
        ws.merge_cells("A2:G2")
        a2 = ws["A2"]
        a2.value = "All figures are calculated directly from the detail tabs, so this panel always reflects current data."
        a2.font = Font(name=self.font_family, size=9, color=self.colors["text_muted"])
        ws.row_dimensions[2].height = 15.75

        # Row 4: Header
        headers = ["Control", f"{file1_name} (File 1)", f"{file2_name} (File 2)", "Result"]
        ws.row_dimensions[4].height = 19.5
        navy_fill = PatternFill("solid", fgColor=self.colors["primary"])
        border = self._thin_border()

        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = Font(name=self.font_family, size=10, bold=True, color=self.colors["header_fg"])
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Rows 5-9: Control checks with dynamic cross-tab formulas
        zebra_fill = PatternFill("solid", fgColor=self.colors["neutral_bg"])
        ctrl_rows = [
            (
                "Total records in scope",
                f"=COUNTA('03 Matched Records'!C5:C{max_matched_row})+SUM('02 Exceptions'!L5:L{max_exc_row})+COUNTA('{sheet_missing_2}'!A5:A{max_m2_row})",
                f"=COUNTA('03 Matched Records'!C5:C{max_matched_row})+SUM('02 Exceptions'!L5:L{max_exc_row})+COUNTA('{sheet_missing_1}'!A5:A{max_m1_row})",
                '=IF(B5=C5,"Pass","Review")',
            ),
            (
                "Matched records (no differences)",
                f"=COUNTA('03 Matched Records'!C5:C{max_matched_row})",
                f"=COUNTA('03 Matched Records'!C5:C{max_matched_row})",
                '=IF(B6=C6,"Pass","Review")',
            ),
            (
                "Exception records (1+ field mismatch)",
                f"=SUM('02 Exceptions'!L5:L{max_exc_row})",
                f"=SUM('02 Exceptions'!L5:L{max_exc_row})",
                '=IF(B7=0,"Pass","Review Required")',
            ),
            (
                f"Records missing from {file1_name}",
                f"=COUNTA('{sheet_missing_1}'!A5:A{max_m1_row})",
                "—",
                '=IF(B8=0,"Pass","Review Required")',
            ),
            (
                f"Records missing from {file2_name}",
                "—",
                f"=COUNTA('{sheet_missing_2}'!A5:A{max_m2_row})",
                '=IF(C9=0,"Pass","Review Required")',
            ),
        ]

        for idx, (label, f1_val, f2_val, res_form) in enumerate(ctrl_rows, start=5):
            fill = zebra_fill if idx % 2 == 0 else PatternFill(fill_type=None)

            c_lbl = ws.cell(row=idx, column=1, value=label)
            c_lbl.font = Font(name=self.font_family, size=9)
            c_lbl.fill = fill
            c_lbl.alignment = Alignment(horizontal="left", vertical="center")
            c_lbl.border = border

            c_f1 = ws.cell(row=idx, column=2, value=f1_val)
            c_f1.font = Font(name=self.font_family, size=9)
            c_f1.fill = fill
            c_f1.alignment = Alignment(horizontal="center", vertical="center")
            c_f1.border = border

            c_f2 = ws.cell(row=idx, column=3, value=f2_val)
            c_f2.font = Font(name=self.font_family, size=9)
            c_f2.fill = fill
            c_f2.alignment = Alignment(horizontal="center", vertical="center")
            c_f2.border = border

            c_res = ws.cell(row=idx, column=4, value=res_form)
            c_res.font = Font(name=self.font_family, size=9, bold=True)
            c_res.fill = fill
            c_res.alignment = Alignment(horizontal="center", vertical="center")
            c_res.border = border

        # Methodology section
        ws.cell(row=11, column=1, value="Methodology").font = Font(
            name=self.font_family, size=10, bold=True, color=self.colors["primary"]
        )

        methodology = [
            f"• Matching key: {key_label} (case-insensitive, numeric-normalized).",
            "• 'Total records in scope' = Matched + Exception + records missing from the opposite file.",
            "• An 'Exception record' has 1 or more field-level mismatches — see tab '02 Exceptions' for detail.",
            f"• Source files: {file1_name} vs {file2_name}.",
        ]
        for i, m_text in enumerate(methodology, start=12):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
            c = ws.cell(row=i, column=1, value=m_text)
            c.font = Font(name=self.font_family, size=9, color=self.colors["text_muted"])

        # Conditional Formatting on D5:D9
        green_fill = PatternFill(start_color=self.colors["pass_bg"], end_color=self.colors["pass_bg"], fill_type="solid")
        green_font = Font(name=self.font_family, size=9, bold=True, color=self.colors["pass"])
        red_fill = PatternFill(start_color=self.colors["fail_bg"], end_color=self.colors["fail_bg"], fill_type="solid")
        red_font = Font(name=self.font_family, size=9, bold=True, color=self.colors["exception"])

        ws.conditional_formatting.add("D5:D9", CellIsRule(operator="equal", formula=['"Pass"'], fill=green_fill, font=green_font))
        ws.conditional_formatting.add("D5:D9", CellIsRule(operator="equal", formula=['"Review Required"'], fill=red_fill, font=red_font))
        ws.conditional_formatting.add("D5:D9", CellIsRule(operator="equal", formula=['"Review"'], fill=red_fill, font=red_font))

        # Column widths
        ws.column_dimensions["A"].width = 34.0
        ws.column_dimensions["B"].width = 20.0
        ws.column_dimensions["C"].width = 20.0
        ws.column_dimensions["D"].width = 16.0
        ws.column_dimensions["E"].width = 6.0


def generate_enterprise_report(data: dict, config: dict, output_path: Path) -> None:
    report_config = ReportConfig(**config)
    reporter = UniversalReporter(data, report_config, output_path)
    reporter.generate()
