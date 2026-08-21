import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from app.api.deps import get_session_id
from app.database.session import get_db
from app.models.job import ReconciliationJob
from app.models.report import Report
from app.storage import get_storage
from pydantic import BaseModel
import tempfile
import json
from app.reconciliation_engine.universal_reporter import generate_enterprise_report


router = APIRouter(prefix="/reports", tags=["reports"])


def _job_report(db: Session, job_id: str, session_id: str) -> tuple[ReconciliationJob, Report]:
    job = db.query(ReconciliationJob).filter(ReconciliationJob.id == job_id, ReconciliationJob.session_id == session_id).first()
    if not job or not job.report_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report not available")
    report = db.query(Report).filter(Report.id == job.report_id, Report.session_id == session_id).first()
    if not report:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report not found")
    return job, report


def _preview_sheet_name(path: Path, category: str) -> str:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        names = workbook.sheetnames
    finally:
        workbook.close()

    if not names:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Workbook is empty")

    if category == "discrepancies":
        for name in names:
            if "02" in name or "Exception" in name or "Discrepanc" in name:
                return name
        return names[0]
    elif category == "only_file_1":
        for name in names:
            if "05" in name:
                return name
        for name in names:
            if "Missing" in name:
                return name
        return names[min(2, len(names) - 1)]
    elif category == "only_file_2":
        for name in names:
            if "04" in name:
                return name
        for name in names:
            if "Missing" in name:
                return name
        return names[min(1, len(names) - 1)]
    elif category == "review":
        for name in names:
            if "06" in name or "Field Difference" in name or "Review" in name:
                return name
        return names[min(3, len(names) - 1)]

    index_by_category = {"discrepancies": 0, "only_file_1": 1, "only_file_2": 2, "review": 3}
    if category in index_by_category and index_by_category[category] < len(names):
        return names[index_by_category[category]]
    return names[0]



def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if pd.isna(value):
        return None
    return value

class ReportCustomConfig(BaseModel):
    include_summary: bool = True
    include_exceptions: bool = True
    include_matched: bool = True
    include_missing_file_1: bool = True
    include_missing_file_2: bool = True
    include_field_differences: bool = True
    include_controls: bool = True
    date_format: str = "YYYY-MM-DD"
    number_format: str = "#,##0.00"


@router.get("/{report_id}/download")
def download_report(
    report_id: str,
    db: Session = Depends(get_db),
    session_id: str = Depends(get_session_id),
) -> FileResponse:
    report = db.query(Report).filter(Report.id == report_id, Report.session_id == session_id).first()
    if not report:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Report not found")
    path = get_storage().resolve_path(report.storage_path)
    return FileResponse(path, filename=report.filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/job/{job_id}/download")
def download_job_report(
    job_id: str,
    db: Session = Depends(get_db),
    session_id: str = Depends(get_session_id),
) -> FileResponse:
    _, report = _job_report(db, job_id, session_id)
    path = get_storage().resolve_path(report.storage_path)
    return FileResponse(path, filename=report.filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/job/{job_id}/download_custom")
def download_custom_report(
    job_id: str,
    config: ReportCustomConfig,
    db: Session = Depends(get_db),
    session_id: str = Depends(get_session_id),
) -> FileResponse:
    _, report = _job_report(db, job_id, session_id)
    storage = get_storage()
    path = storage.resolve_path(report.storage_path)
    raw_path = path.with_name(f"{path.stem}_data.json")
    
    if not raw_path.exists():
        # Fallback to existing Excel file if raw data is lost
        return FileResponse(path, filename=report.filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
    with open(raw_path, "r") as f:
        universal_data = json.load(f)
        
    # Generate new temp file
    from fastapi.background import BackgroundTasks
    import tempfile
    
    fd, temp_path_str = tempfile.mkstemp(suffix=".xlsx", prefix="recliq_custom_")
    import os
    os.close(fd)
    temp_path = Path(temp_path_str)
    
    try:
        generate_enterprise_report(universal_data, config.model_dump(), temp_path)
    except Exception as e:
        if temp_path.exists():
            temp_path.unlink()
        raise HTTPException(status_code=500, detail=str(e))
        
    # We will let FileResponse handle it and optionally delete after? Wait, FileResponse doesn't delete automatically.
    # FastAPI background task can delete it.
    from starlette.background import BackgroundTask
    return FileResponse(
        temp_path, 
        filename=f"Custom_{report.filename}", 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(lambda: temp_path.unlink(missing_ok=True) if temp_path.exists() else None)
    )


@router.get("/job/{job_id}/summary")
def job_report_summary(job_id: str, db: Session = Depends(get_db), session_id: str = Depends(get_session_id)) -> dict[str, Any]:
    _, report = _job_report(db, job_id, session_id)
    try:
        return json.loads(report.summary_json or "{}")
    except json.JSONDecodeError:
        return {}


@router.get("/job/{job_id}/preview")
def job_report_preview(
    job_id: str,
    category: str = Query(default="discrepancies"),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=25, ge=1, le=25),
    db: Session = Depends(get_db),
    session_id: str = Depends(get_session_id),
) -> dict[str, Any]:
    _, report = _job_report(db, job_id, session_id)
    path = get_storage().resolve_path(report.storage_path)
    sheet_name = _preview_sheet_name(path, category)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    header_row_idx = 0
    try:
        ws = workbook[sheet_name]
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            non_null_count = sum(1 for v in row if v is not None and str(v).strip() != "")
            if non_null_count >= 3:
                header_row_idx = r_idx
                break
            if r_idx > 10:
                break
        total_rows = max(0, ws.max_row - (header_row_idx + 1))
    finally:
        workbook.close()

    skip_range = range(header_row_idx + 1, header_row_idx + 1 + offset) if offset > 0 else None
    frame = pd.read_excel(path, sheet_name=sheet_name, header=header_row_idx, skiprows=skip_range, nrows=limit)
    records = [
        {str(column): _json_value(value) for column, value in row.items()}
        for row in frame.to_dict(orient="records")
    ]
    return {
        "category": category,
        "sheet_name": sheet_name,
        "columns": [str(column) for column in frame.columns],
        "rows": records,
        "total_rows": total_rows,
        "offset": offset,
        "limit": limit,
    }
